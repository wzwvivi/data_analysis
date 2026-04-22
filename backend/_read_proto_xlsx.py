# -*- coding: utf-8 -*-
"""读取桌面《协议版本管理.xlsx》为设备树修正做参考。"""
from openpyxl import load_workbook

path = r"/tmp/协议版本管理.xlsx"
wb = load_workbook(path, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"=== sheet: {sn} ({ws.max_row} rows, {ws.max_column} cols) ===")
    for row in ws.iter_rows(values_only=True):
        cells = ["" if v is None else str(v) for v in row]
        print("  " + " | ".join(cells))
    print()
