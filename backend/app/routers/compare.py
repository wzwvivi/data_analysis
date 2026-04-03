# -*- coding: utf-8 -*-
"""双交换机数据比对路由"""
import io
import shutil
import time
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime

from ..database import get_db
from ..config import UPLOAD_DIR, ALLOWED_EXTENSIONS
from ..services import CompareService

router = APIRouter(prefix="/api/compare", tags=["数据比对"])


# ========== Pydantic Schemas ==========

class CompareTaskResponse(BaseModel):
    """比对任务响应"""
    id: int
    filename_1: str
    filename_2: str
    protocol_version_id: int
    status: str
    progress: int
    error_message: Optional[str] = None
    
    # 检查1
    switch1_first_ts: Optional[float] = None
    switch2_first_ts: Optional[float] = None
    time_diff_ms: Optional[float] = None
    sync_result: Optional[str] = None
    
    # 检查2
    expected_port_count: int
    both_present_count: int
    missing_count: int
    
    # 检查3
    periodic_port_count: int
    ports_with_gaps: int
    total_gap_count: int
    
    # 检查4
    jitter_threshold_pct: float
    timing_checked_port_count: int
    timing_pass_count: int
    timing_warning_count: int
    timing_fail_count: int
    
    overall_result: Optional[str] = None
    created_at: datetime
    completed_at: Optional[datetime] = None
    
    class Config:
        from_attributes = True


class CompareTaskListResponse(BaseModel):
    """比对任务列表响应"""
    total: int
    items: List[CompareTaskResponse]


class ComparePortResultResponse(BaseModel):
    """端口比对结果响应"""
    id: int
    port_number: int
    source_device: Optional[str] = None
    message_name: Optional[str] = None
    period_ms: Optional[float] = None
    is_periodic: bool
    
    in_switch1: bool
    in_switch2: bool
    switch1_count: int
    switch2_count: int
    switch1_first_ts: Optional[float] = None
    switch1_last_ts: Optional[float] = None
    switch2_first_ts: Optional[float] = None
    switch2_last_ts: Optional[float] = None
    count_diff: int
    
    gap_count_switch1: int
    gap_count_switch2: int
    
    result: Optional[str] = None
    detail: Optional[str] = None
    
    class Config:
        from_attributes = True


class ComparePortResultListResponse(BaseModel):
    """端口比对结果列表响应"""
    total: int
    items: List[ComparePortResultResponse]


class CompareGapRecordResponse(BaseModel):
    """丢包记录响应"""
    id: int
    port_number: int
    switch_index: int
    gap_start_ts: float
    gap_end_ts: float
    gap_duration_ms: float
    expected_period_ms: float
    estimated_missing_packets: Optional[int] = None
    
    class Config:
        from_attributes = True


class CompareGapRecordListResponse(BaseModel):
    """丢包记录列表响应"""
    total: int
    items: List[CompareGapRecordResponse]


class ComparePortTimingResponse(BaseModel):
    """端口周期正确性与抖动分析响应"""
    id: int
    port_number: int
    switch_index: int
    source_device: Optional[str] = None
    message_name: Optional[str] = None
    expected_period_ms: float
    packet_count: int
    total_intervals: int
    actual_mean_interval_ms: Optional[float] = None
    actual_median_interval_ms: Optional[float] = None
    actual_std_interval_ms: Optional[float] = None
    actual_min_interval_ms: Optional[float] = None
    actual_max_interval_ms: Optional[float] = None
    jitter_pct: Optional[float] = None
    within_threshold_count: int
    compliance_rate_pct: Optional[float] = None
    result: Optional[str] = None
    detail: Optional[str] = None
    
    class Config:
        from_attributes = True


class ComparePortTimingListResponse(BaseModel):
    """端口周期分析结果列表响应"""
    total: int
    items: List[ComparePortTimingResponse]


# ========== API Endpoints ==========

@router.post("/upload")
async def upload_and_compare(
    background_tasks: BackgroundTasks,
    file_1: UploadFile = File(..., description="交换机1的pcap/pcapng文件"),
    file_2: UploadFile = File(..., description="交换机2的pcap/pcapng文件"),
    protocol_version_id: int = Form(..., description="网络配置版本ID"),
    jitter_threshold_pct: float = Form(10.0, description="抖动阈值百分比，默认10%"),
    db: AsyncSession = Depends(get_db)
):
    """上传两个文件并创建比对任务"""
    
    # 先做文件名清洗，防止路径注入
    safe_name_1 = Path(file_1.filename or "").name
    safe_name_2 = Path(file_2.filename or "").name
    if safe_name_1 in {"", ".", ".."} or safe_name_2 in {"", ".", ".."}:
        raise HTTPException(status_code=400, detail="文件名无效")

    # 验证文件类型
    for file, label in [(file_1, "文件1"), (file_2, "文件2")]:
        file_ext = Path(file.filename or "").suffix.lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"{label}不支持的文件类型，只支持: {', '.join(ALLOWED_EXTENSIONS)}"
            )
    
    # 验证网络配置版本
    service = CompareService(db)
    version = await service.protocol_service.get_version(protocol_version_id)
    if not version:
        raise HTTPException(status_code=400, detail="网络配置版本不存在")
    
    # 保存文件
    upload_path = UPLOAD_DIR / "compare"
    upload_path.mkdir(parents=True, exist_ok=True)
    
    timestamp = int(time.time() * 1000)
    safe_filename_1 = f"{timestamp}_switch1_{safe_name_1}"
    safe_filename_2 = f"{timestamp}_switch2_{safe_name_2}"
    
    file_path_1 = upload_path / safe_filename_1
    file_path_2 = upload_path / safe_filename_2
    
    with open(file_path_1, "wb") as f:
        shutil.copyfileobj(file_1.file, f, length=1024*1024)
    
    with open(file_path_2, "wb") as f:
        shutil.copyfileobj(file_2.file, f, length=1024*1024)
    
    # 创建比对任务
    task = await service.create_task(
        filename_1=safe_name_1,
        filename_2=safe_name_2,
        file_path_1=str(file_path_1),
        file_path_2=str(file_path_2),
        protocol_version_id=protocol_version_id,
        jitter_threshold_pct=jitter_threshold_pct
    )
    
    # 后台执行比对
    background_tasks.add_task(run_compare_task, task.id)
    
    return {
        "success": True,
        "task_id": task.id,
        "message": "文件上传成功，比对任务已创建"
    }


async def run_compare_task(task_id: int):
    """后台运行比对任务"""
    import traceback
    print(f"[比对任务] 开始执行任务 {task_id}")
    
    try:
        from ..database import async_session
        
        async with async_session() as db:
            service = CompareService(db)
            result = await service.run_compare(task_id)
            print(f"[比对任务] 任务 {task_id} 完成，结果: {result}")
    except Exception as e:
        print(f"[比对任务] 任务 {task_id} 失败: {e}")
        traceback.print_exc()


@router.get("/tasks", response_model=CompareTaskListResponse)
async def list_tasks(
    page: int = 1,
    page_size: int = 20,
    db: AsyncSession = Depends(get_db)
):
    """获取比对任务列表"""
    service = CompareService(db)
    offset = (page - 1) * page_size
    tasks, total = await service.get_tasks(limit=page_size, offset=offset)
    
    items = [
        CompareTaskResponse(
            id=t.id,
            filename_1=t.filename_1,
            filename_2=t.filename_2,
            protocol_version_id=t.protocol_version_id,
            status=t.status,
            progress=t.progress or 0,
            error_message=t.error_message,
            switch1_first_ts=t.switch1_first_ts,
            switch2_first_ts=t.switch2_first_ts,
            time_diff_ms=t.time_diff_ms,
            sync_result=t.sync_result,
            expected_port_count=t.expected_port_count or 0,
            both_present_count=t.both_present_count or 0,
            missing_count=t.missing_count or 0,
            periodic_port_count=t.periodic_port_count or 0,
            ports_with_gaps=t.ports_with_gaps or 0,
            total_gap_count=t.total_gap_count or 0,
            jitter_threshold_pct=t.jitter_threshold_pct or 10.0,
            timing_checked_port_count=t.timing_checked_port_count or 0,
            timing_pass_count=t.timing_pass_count or 0,
            timing_warning_count=t.timing_warning_count or 0,
            timing_fail_count=t.timing_fail_count or 0,
            overall_result=t.overall_result,
            created_at=t.created_at,
            completed_at=t.completed_at
        )
        for t in tasks
    ]
    
    return CompareTaskListResponse(total=total, items=items)


@router.get("/tasks/{task_id}", response_model=CompareTaskResponse)
async def get_task(
    task_id: int,
    db: AsyncSession = Depends(get_db)
):
    """获取比对任务详情"""
    service = CompareService(db)
    task = await service.get_task(task_id)
    
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    return CompareTaskResponse(
        id=task.id,
        filename_1=task.filename_1,
        filename_2=task.filename_2,
        protocol_version_id=task.protocol_version_id,
        status=task.status,
        progress=task.progress or 0,
        error_message=task.error_message,
        switch1_first_ts=task.switch1_first_ts,
        switch2_first_ts=task.switch2_first_ts,
        time_diff_ms=task.time_diff_ms,
        sync_result=task.sync_result,
        expected_port_count=task.expected_port_count or 0,
        both_present_count=task.both_present_count or 0,
        missing_count=task.missing_count or 0,
        periodic_port_count=task.periodic_port_count or 0,
        ports_with_gaps=task.ports_with_gaps or 0,
        total_gap_count=task.total_gap_count or 0,
        jitter_threshold_pct=task.jitter_threshold_pct or 10.0,
        timing_checked_port_count=task.timing_checked_port_count or 0,
        timing_pass_count=task.timing_pass_count or 0,
        timing_warning_count=task.timing_warning_count or 0,
        timing_fail_count=task.timing_fail_count or 0,
        overall_result=task.overall_result,
        created_at=task.created_at,
        completed_at=task.completed_at
    )


@router.get("/tasks/{task_id}/ports", response_model=ComparePortResultListResponse)
async def get_port_results(
    task_id: int,
    db: AsyncSession = Depends(get_db)
):
    """获取端口比对结果列表"""
    service = CompareService(db)
    
    # 验证任务存在
    task = await service.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    port_results = await service.get_port_results(task_id)
    
    items = [
        ComparePortResultResponse(
            id=r.id,
            port_number=r.port_number,
            source_device=r.source_device,
            message_name=r.message_name,
            period_ms=r.period_ms,
            is_periodic=r.is_periodic,
            in_switch1=r.in_switch1,
            in_switch2=r.in_switch2,
            switch1_count=r.switch1_count,
            switch2_count=r.switch2_count,
            switch1_first_ts=r.switch1_first_ts,
            switch1_last_ts=r.switch1_last_ts,
            switch2_first_ts=r.switch2_first_ts,
            switch2_last_ts=r.switch2_last_ts,
            count_diff=r.count_diff,
            gap_count_switch1=r.gap_count_switch1,
            gap_count_switch2=r.gap_count_switch2,
            result=r.result,
            detail=r.detail
        )
        for r in port_results
    ]
    
    return ComparePortResultListResponse(total=len(items), items=items)


@router.get("/tasks/{task_id}/gaps", response_model=CompareGapRecordListResponse)
async def get_gap_records(
    task_id: int,
    port: int = None,
    db: AsyncSession = Depends(get_db)
):
    """获取丢包记录列表"""
    service = CompareService(db)
    
    # 验证任务存在
    task = await service.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    gap_records = await service.get_gap_records(task_id, port_number=port)
    
    items = [
        CompareGapRecordResponse(
            id=r.id,
            port_number=r.port_number,
            switch_index=r.switch_index,
            gap_start_ts=r.gap_start_ts,
            gap_end_ts=r.gap_end_ts,
            gap_duration_ms=r.gap_duration_ms,
            expected_period_ms=r.expected_period_ms,
            estimated_missing_packets=r.estimated_missing_packets
        )
        for r in gap_records
    ]
    
    return CompareGapRecordListResponse(total=len(items), items=items)


@router.get("/tasks/{task_id}/timing", response_model=ComparePortTimingListResponse)
async def get_timing_results(
    task_id: int,
    port: int = None,
    switch: int = None,
    db: AsyncSession = Depends(get_db)
):
    """获取端口周期正确性与抖动分析结果"""
    service = CompareService(db)
    
    task = await service.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    
    timing_results = await service.get_timing_results(task_id, port_number=port, switch_index=switch)
    
    items = [
        ComparePortTimingResponse(
            id=r.id,
            port_number=r.port_number,
            switch_index=r.switch_index,
            source_device=r.source_device,
            message_name=r.message_name,
            expected_period_ms=r.expected_period_ms,
            packet_count=r.packet_count,
            total_intervals=r.total_intervals,
            actual_mean_interval_ms=r.actual_mean_interval_ms,
            actual_median_interval_ms=r.actual_median_interval_ms,
            actual_std_interval_ms=r.actual_std_interval_ms,
            actual_min_interval_ms=r.actual_min_interval_ms,
            actual_max_interval_ms=r.actual_max_interval_ms,
            jitter_pct=r.jitter_pct,
            within_threshold_count=r.within_threshold_count,
            compliance_rate_pct=r.compliance_rate_pct,
            result=r.result,
            detail=r.detail
        )
        for r in timing_results
    ]
    
    return ComparePortTimingListResponse(total=len(items), items=items)


@router.get("/tasks/{task_id}/export")
async def export_report(
    task_id: int,
    db: AsyncSession = Depends(get_db)
):
    """导出比对报告为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    service = CompareService(db)
    task = await service.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    if task.status != "completed":
        raise HTTPException(status_code=400, detail="任务尚未完成，无法导出")
    
    port_results = await service.get_port_results(task_id)
    gap_records = await service.get_gap_records(task_id)
    timing_results = await service.get_timing_results(task_id)
    
    wb = Workbook()
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fail_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    
    def result_fill(val):
        if val == "pass": return pass_fill
        if val == "warning": return warning_fill
        if val == "fail": return fail_fill
        return None
    
    def result_text(val):
        if val == "pass": return "通过"
        if val == "warning": return "警告"
        if val == "fail": return "失败"
        return str(val or "未知")
    
    def write_header(ws, headers):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
    
    def style_cell(cell, result=None):
        cell.border = thin_border
        cell.alignment = cell_align
        fill = result_fill(result)
        if fill:
            cell.fill = fill
    
    def format_ts(ts):
        if not ts:
            return "-"
        from datetime import datetime as dt, timezone, timedelta
        d = dt.fromtimestamp(ts, tz=timezone(timedelta(hours=8)))
        return d.strftime("%Y-%m-%d %H:%M:%S.") + f"{d.microsecond // 1000:03d}"
    
    # === Sheet 1: 比对概览 ===
    ws1 = wb.active
    ws1.title = "比对概览"
    overview_data = [
        ["比对任务ID", task.id],
        ["交换机1文件", task.filename_1],
        ["交换机2文件", task.filename_2],
        ["创建时间", str(task.created_at)],
        ["完成时间", str(task.completed_at or "")],
        ["综合结论", result_text(task.overall_result)],
        [""],
        ["检查1: 记录时间同步性"],
        ["交换机1首包时间", format_ts(task.switch1_first_ts)],
        ["交换机2首包时间", format_ts(task.switch2_first_ts)],
        ["时间差(ms)", f"{task.time_diff_ms:.3f}" if task.time_diff_ms is not None else "-"],
        ["同步结果", result_text(task.sync_result)],
        [""],
        ["检查2: 端口覆盖完整性"],
        ["网络配置端口总数", task.expected_port_count or 0],
        ["两边都有数据", task.both_present_count or 0],
        ["至少一边缺失", task.missing_count or 0],
        [""],
        ["检查3: 周期端口数据连续性"],
        ["周期类端口总数", task.periodic_port_count or 0],
        ["存在丢包的端口", task.ports_with_gaps or 0],
        ["总丢包段数", task.total_gap_count or 0],
        [""],
        ["检查4: 端口周期正确性与抖动分析"],
        ["抖动阈值(%)", task.jitter_threshold_pct or 10.0],
        ["检查端口数", task.timing_checked_port_count or 0],
        ["通过", task.timing_pass_count or 0],
        ["警告", task.timing_warning_count or 0],
        ["失败", task.timing_fail_count or 0],
    ]
    
    title_font = Font(bold=True, size=12, color="1F4E79")
    for row_idx, row_data in enumerate(overview_data, 1):
        if len(row_data) == 1 and isinstance(row_data[0], str) and row_data[0].startswith("检查"):
            cell = ws1.cell(row=row_idx, column=1, value=row_data[0])
            cell.font = title_font
        elif len(row_data) == 2:
            ws1.cell(row=row_idx, column=1, value=row_data[0]).font = Font(bold=True)
            ws1.cell(row=row_idx, column=2, value=row_data[1])
    
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 40
    
    # === Sheet 2: 端口覆盖完整性 ===
    ws2 = wb.create_sheet("端口覆盖完整性")
    port_headers = ["端口号", "源设备", "消息名称", "周期(ms)", "交换机1包数", "交换机2包数",
                    "包数差", "交换机1丢包段", "交换机2丢包段", "结果", "说明"]
    write_header(ws2, port_headers)
    
    for row_idx, r in enumerate(port_results, 2):
        values = [
            r.port_number,
            r.source_device or "",
            r.message_name or "",
            f"{r.period_ms:.1f}" if r.period_ms else "-",
            r.switch1_count if r.in_switch1 else "该端口未出现",
            r.switch2_count if r.in_switch2 else "该端口未出现",
            r.count_diff,
            r.gap_count_switch1 or 0,
            r.gap_count_switch2 or 0,
            result_text(r.result),
            r.detail or "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            style_cell(cell, result=r.result)
    
    for col, w in [(1, 10), (2, 18), (3, 22), (4, 12), (5, 14), (6, 14),
                   (7, 10), (8, 14), (9, 14), (10, 10), (11, 30)]:
        ws2.column_dimensions[chr(64 + col)].width = w
    
    # === Sheet 3: 丢包记录 ===
    ws3 = wb.create_sheet("丢包记录")
    gap_headers = ["端口号", "交换机", "丢包起始时间", "丢包结束时间", "间隔时长(ms)",
                   "预期周期(ms)", "预估缺失包数"]
    write_header(ws3, gap_headers)
    
    for row_idx, r in enumerate(gap_records, 2):
        values = [
            r.port_number,
            f"交换机{r.switch_index}",
            format_ts(r.gap_start_ts),
            format_ts(r.gap_end_ts),
            f"{r.gap_duration_ms:.2f}",
            f"{r.expected_period_ms:.1f}",
            r.estimated_missing_packets or 0,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            style_cell(cell)
    
    for col, w in [(1, 10), (2, 12), (3, 24), (4, 24), (5, 14), (6, 14), (7, 14)]:
        ws3.column_dimensions[chr(64 + col)].width = w
    
    if not gap_records:
        ws3.cell(row=2, column=1, value="未检测到丢包").font = Font(italic=True, color="808080")
    
    # === Sheet 4 & 5: 周期分析 (交换机1 / 交换机2) ===
    timing_headers = ["端口号", "源设备", "消息名称", "预期周期(ms)", "包数",
                      "实际均值(ms)", "中位数(ms)", "标准差(ms)", "最小间隔(ms)", "最大间隔(ms)",
                      "抖动%", "达标率%", "结果", "说明"]
    
    for sw_idx, sw_label in [(1, "交换机1"), (2, "交换机2")]:
        ws = wb.create_sheet(f"周期分析-{sw_label}")
        write_header(ws, timing_headers)
        
        sw_data = [r for r in timing_results if r.switch_index == sw_idx]
        for row_idx, r in enumerate(sw_data, 2):
            values = [
                r.port_number,
                r.source_device or "",
                r.message_name or "",
                f"{r.expected_period_ms:.1f}",
                r.packet_count,
                f"{r.actual_mean_interval_ms:.2f}" if r.actual_mean_interval_ms is not None else "-",
                f"{r.actual_median_interval_ms:.2f}" if r.actual_median_interval_ms is not None else "-",
                f"{r.actual_std_interval_ms:.2f}" if r.actual_std_interval_ms is not None else "-",
                f"{r.actual_min_interval_ms:.2f}" if r.actual_min_interval_ms is not None else "-",
                f"{r.actual_max_interval_ms:.2f}" if r.actual_max_interval_ms is not None else "-",
                f"{r.jitter_pct:.1f}%" if r.jitter_pct is not None else "-",
                f"{r.compliance_rate_pct:.1f}%" if r.compliance_rate_pct is not None else "-",
                result_text(r.result),
                r.detail or "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                style_cell(cell, result=r.result)
        
        for col, w in [(1, 10), (2, 18), (3, 22), (4, 14), (5, 8),
                       (6, 14), (7, 14), (8, 14), (9, 14), (10, 14),
                       (11, 10), (12, 10), (13, 10), (14, 24)]:
            ws.column_dimensions[chr(64 + col)].width = w
        
        if not sw_data:
            ws.cell(row=2, column=1, value="无数据").font = Font(italic=True, color="808080")
    
    # 冻结首行
    for ws in wb.worksheets[1:]:
        ws.freeze_panes = "A2"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"compare_report_task{task_id}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
