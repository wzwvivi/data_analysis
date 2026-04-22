# -*- coding: utf-8 -*-
"""构型管理服务层

提供：
  - Device / AircraftConfiguration / SoftwareConfiguration 的查询与增删改；
  - 软件构型 Excel 一键导入：解析《软件编号（首飞构型定义A版）.xlsx》风格的 3 个 sheet，
    upsert 设备库并按列创建 SoftwareConfiguration + SoftwareConfigurationEntry。

Excel 列约定（1-based 为对照便于阅读，代码里按 0-based 索引）：
  0  序号
  1  软件归属信息       -> team
  2  EATA 章节号-系统名称 -> eata_chapter
  3  设备中文名称       -> device_cn_name
  4  设备 DM 号         -> device_dm_number
  5  软件中文名称       -> software_cn_name
  6  软件等级           -> software_level
  7  是否为显控计算机驻留的软件 -> is_cds_resident
  8  是否外场可加载的软件 -> is_field_loadable
  9  是否为自研软件     -> is_proprietary
  10 软件供应商         -> supplier
  11 是否新研软件       -> is_new_dev
  12 是否有软件         -> has_software
  13+ 交替：软件编号（...） / 较*更改说明
       - 第 1 个构型列右侧可能不跟 "较..." 列；
       - 任何以"较"开头、含"说明"的表头都视为"左侧最近构型列"的 change_note 列。
"""
from __future__ import annotations

import logging
import re
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..models import (
    AircraftConfigDeviceProtocolLink,
    AircraftConfiguration,
    Device,
    DeviceProtocolVersion,
    ProtocolVersion,
    SW_CONFIG_SOURCE_EXCEL,
    SoftwareConfiguration,
    SoftwareConfigurationEntry,
)

logger = logging.getLogger(__name__)


# ── 基础工具 ───────────────────────────────────────────────────────────────
_TRUE_TOKENS = {"是", "Y", "YES", "TRUE", "1", "有"}
_FALSE_TOKENS = {"否", "N", "NO", "FALSE", "0", "无"}


def _coerce_bool(value: Any) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().upper()
    if not s or s in {"/", "-", "N/A", "NA", "NONE"}:
        return None
    if s in _TRUE_TOKENS:
        return True
    if s in _FALSE_TOKENS:
        return False
    return None


def _coerce_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


_DATE_PATTERN = re.compile(r"(20\d{2})[.\-/年](\d{1,2})[.\-/月](\d{1,2})")


def _extract_snapshot_date(config_name: str) -> Optional[date]:
    """从 '软件编号（机上试验2026.04.09）' 之类的文本里抓一个 yyyy.mm.dd。"""
    if not config_name:
        return None
    m = _DATE_PATTERN.search(config_name)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


_CONFIG_HEADER_RE = re.compile(r"^\s*软件编号\s*[（(](.+?)[）)]\s*$")


def _parse_config_header(header: Optional[str]) -> Optional[str]:
    """'软件编号（机上试验2026.01.10）' -> '机上试验2026.01.10'；不是构型列时返回 None。"""
    if not header:
        return None
    m = _CONFIG_HEADER_RE.match(str(header))
    if not m:
        return None
    return m.group(1).strip() or None


def _is_change_note_header(header: Optional[str]) -> bool:
    if not header:
        return False
    s = str(header).strip()
    if not s:
        return False
    # 表头里 `较xx更改说明` / `较上次更改说明` / `较首飞构型定义更改说明`
    return s.startswith("较") and ("说明" in s or s.endswith("更改"))


# ── Excel 导入核心 ─────────────────────────────────────────────────────────
_DEVICE_COL_COUNT = 13  # 前 13 列为设备元数据

_DEVICE_META_COLS = [
    ("team", 1),
    ("eata_chapter", 2),
    ("device_cn_name", 3),
    ("device_dm_number", 4),
    ("software_cn_name", 5),
    ("software_level", 6),
    ("is_cds_resident", 7),
    ("is_field_loadable", 8),
    ("is_proprietary", 9),
    ("supplier", 10),
    ("is_new_dev", 11),
    ("has_software", 12),
]

_BOOL_FIELDS = {
    "is_cds_resident",
    "is_field_loadable",
    "is_proprietary",
    "is_new_dev",
    "has_software",
}


class ExcelImportSummary(dict):
    """便于 pydantic 序列化的 dict 容器。"""


def _row_to_device_payload(row: Sequence[Any], *, fallback_team: str) -> Optional[Dict[str, Any]]:
    payload: Dict[str, Any] = {}
    for field, idx in _DEVICE_META_COLS:
        raw = row[idx] if idx < len(row) else None
        if field in _BOOL_FIELDS:
            payload[field] = _coerce_bool(raw)
        else:
            payload[field] = _coerce_str(raw)
    if not payload.get("team"):
        payload["team"] = fallback_team
    # 关键三元组缺失则视为空行（表头/合并单元格）
    if not (payload.get("team") and payload.get("device_cn_name") and payload.get("software_cn_name")):
        return None
    return payload


async def _upsert_device(db: AsyncSession, payload: Dict[str, Any]) -> Tuple[Device, bool]:
    """按 (team, device_cn_name, software_cn_name) upsert 一条 Device。返回 (row, created?)。"""
    q = await db.execute(
        select(Device).where(
            Device.team == payload["team"],
            Device.device_cn_name == payload["device_cn_name"],
            Device.software_cn_name == payload["software_cn_name"],
        )
    )
    row = q.scalar_one_or_none()
    if row is None:
        row = Device(**payload)
        db.add(row)
        await db.flush()
        return row, True
    # 合并元数据：空值不覆盖已有值（便于第二次导入不丢信息）
    for k, v in payload.items():
        if v is None or v == "":
            continue
        setattr(row, k, v)
    return row, False


def _scan_config_columns(header_row: Sequence[Any]) -> List[Tuple[int, str, Optional[int]]]:
    """遍历表头，返回 [(config_col_idx, config_name, change_note_col_idx_or_None), ...]。"""
    items: List[Tuple[int, str, Optional[int]]] = []
    pending: List[List[Any]] = []  # list of [idx, name, note_idx]
    for idx, cell in enumerate(header_row):
        if idx < _DEVICE_COL_COUNT:
            continue
        header = cell.value if hasattr(cell, "value") else cell
        name = _parse_config_header(header)
        if name:
            pending.append([idx, name, None])
            continue
        if _is_change_note_header(header):
            if pending and pending[-1][2] is None:
                pending[-1][2] = idx
    for idx, name, note_idx in pending:
        items.append((idx, name, note_idx))
    return items


async def import_excel_bytes(
    db: AsyncSession,
    *,
    file_bytes: bytes,
    source_file: str,
    created_by: Optional[str],
    mode: str = "merge",
) -> ExcelImportSummary:
    """解析 bytes，upsert 设备并为每个构型列写入 SoftwareConfiguration + entries。

    mode:
      - 'merge' 默认：软件构型名称已存在则在其条目上做 upsert；
      - 'replace'：软件构型已存在则清空其 entries 后再写入（保留构型本身，id 稳定）。
    """
    import openpyxl  # 懒加载避免启动时成本

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    devices_created = 0
    devices_updated = 0
    configs_created = 0
    configs_updated = 0
    entries_written = 0
    skipped_rows = 0
    warnings: List[str] = []

    # 某个构型可能在多个 sheet 中复现（历史数据极少但可能发生），用 cache 去重
    config_cache: Dict[str, SoftwareConfiguration] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        config_cols = _scan_config_columns(header_row)
        if not config_cols:
            warnings.append(f"sheet={sheet_name} 未识别到任何 '软件编号（...）' 列")
            continue

        # 先处理设备行
        sheet_device_rows: List[Tuple[Dict[str, Any], Device, Sequence[Any]]] = []
        for row_cells in ws.iter_rows(min_row=2):
            row_values = [c.value for c in row_cells]
            payload = _row_to_device_payload(row_values, fallback_team=str(sheet_name))
            if not payload:
                skipped_rows += 1
                continue
            device, created = await _upsert_device(db, payload)
            if created:
                devices_created += 1
            else:
                devices_updated += 1
            sheet_device_rows.append((payload, device, row_values))
        await db.flush()

        # 再按构型列写入 entries
        for col_idx, config_name, note_idx in config_cols:
            cfg = config_cache.get(config_name)
            if cfg is None:
                q = await db.execute(
                    select(SoftwareConfiguration).where(SoftwareConfiguration.name == config_name)
                )
                cfg = q.scalar_one_or_none()
            created_cfg = False
            if cfg is None:
                cfg = SoftwareConfiguration(
                    name=config_name,
                    snapshot_date=_extract_snapshot_date(config_name),
                    source=SW_CONFIG_SOURCE_EXCEL,
                    source_file=source_file,
                    description=None,
                    created_by=created_by,
                )
                db.add(cfg)
                await db.flush()
                created_cfg = True
                configs_created += 1
            else:
                cfg.source = cfg.source or SW_CONFIG_SOURCE_EXCEL
                cfg.source_file = cfg.source_file or source_file
                if not cfg.snapshot_date:
                    cfg.snapshot_date = _extract_snapshot_date(config_name)
                cfg.updated_at = datetime.utcnow()
                configs_updated += 1
                if mode == "replace":
                    await db.execute(
                        delete(SoftwareConfigurationEntry).where(
                            SoftwareConfigurationEntry.software_config_id == cfg.id
                        )
                    )
            config_cache[config_name] = cfg

            for payload, device, row_values in sheet_device_rows:
                version_code = _coerce_str(row_values[col_idx] if col_idx < len(row_values) else None)
                change_note = None
                if note_idx is not None and note_idx < len(row_values):
                    change_note = _coerce_str(row_values[note_idx])
                if version_code is None and change_note is None:
                    continue
                existing_q = await db.execute(
                    select(SoftwareConfigurationEntry).where(
                        SoftwareConfigurationEntry.software_config_id == cfg.id,
                        SoftwareConfigurationEntry.device_id == device.id,
                    )
                )
                entry = existing_q.scalar_one_or_none()
                if entry is None:
                    entry = SoftwareConfigurationEntry(
                        software_config_id=cfg.id,
                        device_id=device.id,
                        software_version_code=version_code,
                        change_note=change_note,
                    )
                    db.add(entry)
                else:
                    if version_code is not None:
                        entry.software_version_code = version_code
                    if change_note is not None:
                        entry.change_note = change_note
                    entry.updated_at = datetime.utcnow()
                entries_written += 1

    await db.commit()
    return ExcelImportSummary(
        devices_created=devices_created,
        devices_updated=devices_updated,
        configs_created=configs_created,
        configs_updated=configs_updated,
        entries_written=entries_written,
        skipped_rows=skipped_rows,
        warnings=warnings,
    )


# ── 查询辅助 ───────────────────────────────────────────────────────────────
async def list_devices(
    db: AsyncSession,
    *,
    team: Optional[str] = None,
    keyword: Optional[str] = None,
) -> List[Device]:
    stmt = select(Device)
    if team:
        stmt = stmt.where(Device.team == team)
    if keyword:
        like = f"%{keyword}%"
        stmt = stmt.where(
            (Device.device_cn_name.like(like))
            | (Device.software_cn_name.like(like))
            | (Device.device_dm_number.like(like))
            | (Device.eata_chapter.like(like))
        )
    stmt = stmt.order_by(Device.team, Device.eata_chapter, Device.device_cn_name, Device.id)
    r = await db.execute(stmt)
    return list(r.scalars().all())


async def list_aircraft_configurations(db: AsyncSession) -> List[AircraftConfiguration]:
    r = await db.execute(
        select(AircraftConfiguration)
        .options(selectinload(AircraftConfiguration.device_protocol_links))
        .order_by(AircraftConfiguration.created_at.desc())
    )
    return list(r.scalars().unique().all())


async def list_software_configurations(db: AsyncSession) -> List[SoftwareConfiguration]:
    r = await db.execute(
        select(SoftwareConfiguration).order_by(
            SoftwareConfiguration.snapshot_date.is_(None),  # 有日期优先
            SoftwareConfiguration.snapshot_date.desc(),
            SoftwareConfiguration.created_at.desc(),
        )
    )
    return list(r.scalars().all())


async def get_aircraft_configuration(
    db: AsyncSession, cfg_id: int
) -> Optional[AircraftConfiguration]:
    r = await db.execute(
        select(AircraftConfiguration)
        .where(AircraftConfiguration.id == cfg_id)
        .options(selectinload(AircraftConfiguration.device_protocol_links))
    )
    return r.scalar_one_or_none()


async def get_software_configuration(
    db: AsyncSession, cfg_id: int
) -> Optional[SoftwareConfiguration]:
    r = await db.execute(
        select(SoftwareConfiguration).where(SoftwareConfiguration.id == cfg_id)
    )
    return r.scalar_one_or_none()


async def list_software_entries(
    db: AsyncSession, cfg_id: int
) -> List[Tuple[SoftwareConfigurationEntry, Device]]:
    r = await db.execute(
        select(SoftwareConfigurationEntry, Device)
        .join(Device, SoftwareConfigurationEntry.device_id == Device.id)
        .where(SoftwareConfigurationEntry.software_config_id == cfg_id)
        .order_by(Device.team, Device.eata_chapter, Device.device_cn_name, Device.id)
    )
    return [(e, d) for e, d in r.all()]


# ── 飞机构型：创建 / 更新 ──────────────────────────────────────────────────
async def _ensure_protocol_version_exists(
    db: AsyncSession, pv_id: Optional[int]
) -> None:
    if pv_id is None:
        return
    r = await db.execute(select(ProtocolVersion.id).where(ProtocolVersion.id == pv_id))
    if r.scalar_one_or_none() is None:
        raise ValueError(f"TSN 协议版本不存在: {pv_id}")


async def _ensure_device_protocol_versions_exist(
    db: AsyncSession, ids: Iterable[int]
) -> List[int]:
    unique_ids = sorted({int(i) for i in ids})
    if not unique_ids:
        return []
    r = await db.execute(
        select(DeviceProtocolVersion.id).where(DeviceProtocolVersion.id.in_(unique_ids))
    )
    existing = {int(x) for x in r.scalars().all()}
    missing = [i for i in unique_ids if i not in existing]
    if missing:
        raise ValueError(f"设备协议版本不存在: {missing}")
    return unique_ids


async def create_aircraft_configuration(
    db: AsyncSession,
    *,
    name: str,
    version: Optional[str],
    description: Optional[str],
    tsn_protocol_version_id: Optional[int],
    device_protocol_version_ids: Optional[List[int]],
    created_by: Optional[str],
) -> AircraftConfiguration:
    name = (name or "").strip()
    if not name:
        raise ValueError("构型名称不能为空")
    await _ensure_protocol_version_exists(db, tsn_protocol_version_id)
    dp_ids = await _ensure_device_protocol_versions_exist(db, device_protocol_version_ids or [])

    existing = await db.execute(
        select(AircraftConfiguration).where(AircraftConfiguration.name == name)
    )
    if existing.scalar_one_or_none():
        raise ValueError(f"飞机构型名称已存在: {name}")

    row = AircraftConfiguration(
        name=name,
        version=(version or None),
        description=(description or None),
        tsn_protocol_version_id=tsn_protocol_version_id,
        created_by=created_by,
    )
    db.add(row)
    await db.flush()
    for dp_id in dp_ids:
        db.add(
            AircraftConfigDeviceProtocolLink(
                aircraft_config_id=row.id,
                device_protocol_version_id=dp_id,
            )
        )
    await db.commit()
    return await get_aircraft_configuration(db, row.id)  # type: ignore[return-value]


async def update_aircraft_configuration(
    db: AsyncSession,
    row: AircraftConfiguration,
    *,
    name: Optional[str] = None,
    version: Optional[str] = None,
    description: Optional[str] = None,
    tsn_protocol_version_id: Optional[int] = None,
    device_protocol_version_ids: Optional[List[int]] = None,
    patch_keys: Optional[set] = None,
) -> AircraftConfiguration:
    patch_keys = patch_keys or set()
    if "name" in patch_keys and name is not None:
        name = name.strip()
        if not name:
            raise ValueError("构型名称不能为空")
        dup = await db.execute(
            select(AircraftConfiguration).where(
                AircraftConfiguration.name == name,
                AircraftConfiguration.id != row.id,
            )
        )
        if dup.scalar_one_or_none():
            raise ValueError(f"飞机构型名称已存在: {name}")
        row.name = name
    if "version" in patch_keys:
        row.version = (version or None)
    if "description" in patch_keys:
        row.description = (description or None)
    if "tsn_protocol_version_id" in patch_keys:
        await _ensure_protocol_version_exists(db, tsn_protocol_version_id)
        row.tsn_protocol_version_id = tsn_protocol_version_id
    if "device_protocol_version_ids" in patch_keys:
        target_ids = await _ensure_device_protocol_versions_exist(db, device_protocol_version_ids or [])
        await db.execute(
            delete(AircraftConfigDeviceProtocolLink).where(
                AircraftConfigDeviceProtocolLink.aircraft_config_id == row.id
            )
        )
        for dp_id in target_ids:
            db.add(
                AircraftConfigDeviceProtocolLink(
                    aircraft_config_id=row.id,
                    device_protocol_version_id=dp_id,
                )
            )
    await db.commit()
    return await get_aircraft_configuration(db, row.id)  # type: ignore[return-value]


async def delete_aircraft_configuration(
    db: AsyncSession, row: AircraftConfiguration
) -> None:
    await db.execute(
        delete(AircraftConfiguration).where(AircraftConfiguration.id == row.id)
    )
    await db.commit()


# ── 软件构型：创建 / 更新 / 条目维护 ───────────────────────────────────────
async def create_software_configuration(
    db: AsyncSession,
    *,
    name: str,
    snapshot_date: Optional[date],
    description: Optional[str],
    created_by: Optional[str],
) -> SoftwareConfiguration:
    name = (name or "").strip()
    if not name:
        raise ValueError("软件构型名称不能为空")
    existing = await db.execute(
        select(SoftwareConfiguration).where(SoftwareConfiguration.name == name)
    )
    if existing.scalar_one_or_none():
        raise ValueError(f"软件构型名称已存在: {name}")
    row = SoftwareConfiguration(
        name=name,
        snapshot_date=snapshot_date or _extract_snapshot_date(name),
        description=(description or None),
        created_by=created_by,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return row


async def update_software_configuration(
    db: AsyncSession,
    row: SoftwareConfiguration,
    *,
    name: Optional[str] = None,
    snapshot_date: Optional[date] = None,
    description: Optional[str] = None,
    patch_keys: Optional[set] = None,
) -> SoftwareConfiguration:
    patch_keys = patch_keys or set()
    if "name" in patch_keys and name is not None:
        name = name.strip()
        if not name:
            raise ValueError("软件构型名称不能为空")
        dup = await db.execute(
            select(SoftwareConfiguration).where(
                SoftwareConfiguration.name == name,
                SoftwareConfiguration.id != row.id,
            )
        )
        if dup.scalar_one_or_none():
            raise ValueError(f"软件构型名称已存在: {name}")
        row.name = name
    if "snapshot_date" in patch_keys:
        row.snapshot_date = snapshot_date
    if "description" in patch_keys:
        row.description = (description or None)
    await db.commit()
    await db.refresh(row)
    return row


async def delete_software_configuration(
    db: AsyncSession, row: SoftwareConfiguration
) -> None:
    await db.execute(
        delete(SoftwareConfiguration).where(SoftwareConfiguration.id == row.id)
    )
    await db.commit()


async def upsert_software_entries(
    db: AsyncSession,
    cfg: SoftwareConfiguration,
    items: List[Dict[str, Any]],
    *,
    replace_all: bool = False,
) -> int:
    """按 items=[{device_id, software_version_code, change_note}] 更新条目。"""
    if replace_all:
        await db.execute(
            delete(SoftwareConfigurationEntry).where(
                SoftwareConfigurationEntry.software_config_id == cfg.id
            )
        )
    count = 0
    for item in items:
        dev_id = int(item.get("device_id"))
        if not dev_id:
            continue
        q = await db.execute(
            select(SoftwareConfigurationEntry).where(
                SoftwareConfigurationEntry.software_config_id == cfg.id,
                SoftwareConfigurationEntry.device_id == dev_id,
            )
        )
        entry = q.scalar_one_or_none()
        version_code = _coerce_str(item.get("software_version_code"))
        change_note = _coerce_str(item.get("change_note"))
        if entry is None:
            # 全部字段都为空就不新增了
            if version_code is None and change_note is None:
                continue
            entry = SoftwareConfigurationEntry(
                software_config_id=cfg.id,
                device_id=dev_id,
                software_version_code=version_code,
                change_note=change_note,
            )
            db.add(entry)
        else:
            entry.software_version_code = version_code
            entry.change_note = change_note
            entry.updated_at = datetime.utcnow()
        count += 1
    await db.commit()
    return count


# ── 设备库：手工维护 ───────────────────────────────────────────────────────
async def create_device(db: AsyncSession, payload: Dict[str, Any]) -> Device:
    payload = dict(payload)
    for k in ("team", "device_cn_name", "software_cn_name"):
        if not payload.get(k):
            raise ValueError(f"缺少必填字段: {k}")
    dup = await db.execute(
        select(Device).where(
            Device.team == payload["team"],
            Device.device_cn_name == payload["device_cn_name"],
            Device.software_cn_name == payload["software_cn_name"],
        )
    )
    if dup.scalar_one_or_none():
        raise ValueError("同 (团队, 设备名, 软件名) 的设备已存在")
    row = Device(**payload)
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return row


async def update_device(
    db: AsyncSession, row: Device, payload: Dict[str, Any]
) -> Device:
    # 若改了业务键，校验唯一
    new_team = payload.get("team", row.team)
    new_cn = payload.get("device_cn_name", row.device_cn_name)
    new_sw = payload.get("software_cn_name", row.software_cn_name)
    if (new_team, new_cn, new_sw) != (row.team, row.device_cn_name, row.software_cn_name):
        dup = await db.execute(
            select(Device).where(
                Device.team == new_team,
                Device.device_cn_name == new_cn,
                Device.software_cn_name == new_sw,
                Device.id != row.id,
            )
        )
        if dup.scalar_one_or_none():
            raise ValueError("同 (团队, 设备名, 软件名) 的设备已存在")
    for k, v in payload.items():
        if hasattr(row, k):
            setattr(row, k, v)
    row.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(row)
    return row


async def delete_device(db: AsyncSession, row: Device) -> None:
    await db.execute(delete(Device).where(Device.id == row.id))
    await db.commit()


async def get_device_by_id(db: AsyncSession, device_id: int) -> Optional[Device]:
    r = await db.execute(select(Device).where(Device.id == device_id))
    return r.scalar_one_or_none()
